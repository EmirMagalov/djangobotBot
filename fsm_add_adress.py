import os
from aiogram import Router,types,F,Bot
from openpyxl import load_workbook
import logging
from aiogram.filters import StateFilter
from aiogram.fsm.state import StatesGroup,State
from aiogram.fsm.context import FSMContext
import kbds
from add_basket_data import *
from aiogram.types import LabeledPrice
from datetime import datetime
from dotenv import load_dotenv
load_dotenv()
fsm_add_adress=Router()
ADMIN=7287871980
class AddData(StatesGroup):
    name=State()
    phone = State()
    adress=State()



@fsm_add_adress.message(F.text=="Отмена")
async def fsm_cancel(message:types.Message, state:FSMContext):
    from private import start

    await message.answer("Покупка отменена",reply_markup=kbds.remove_kb)
    await state.clear()
    await start(message,state)

@fsm_add_adress.message(AddData.name,F.text)
async def fsm_name(message:types.Message, state:FSMContext):
    await state.update_data(name=message.text)
    await message.answer("Укажи номер телефона")
    await state.set_state(AddData.phone)

@fsm_add_adress.message(AddData.phone,F.text)
async def fsm_phone(message:types.Message, state:FSMContext):
    await state.update_data(phone=message.text)
    await message.answer("Укажи свой адрес")
    await state.set_state(AddData.adress)

@fsm_add_adress.message(AddData.adress,F.text)
async def fsm_adress(message:types.Message,bot:Bot, state:FSMContext):
    await state.update_data(adress=message.text)
    await state.clear()
    await fsm_finish(message,bot,state)

async def fsm_finish(message:types.Message,bot:Bot,state:FSMContext):
    data = await state.get_data()
    await message.answer(f'Звать: {data["name"]}\nНомер: {data["phone"]}\nАдресс: {data["adress"]}',reply_markup=kbds.remove_kb)
    user_id = int(message.from_user.id)
    a = await get_basket(user_id)
    total_price = 0
    total_name = []
    for i in a:
        p_name = i["product_name"]
        p_quan = i["quantity"]
        p_price = float(i["product_price"]) * int(p_quan)
        total_name.append(p_name)
        total_price += p_price
    await bot.send_invoice(
        chat_id=user_id,
        title=",".join(total_name),
        description=f"Опалта товаров на сумму {total_price}",
        payload=",".join(total_name),
        provider_token=os.getenv("YOOTOKEN"),
        currency=os.getenv("CURRENCY"),
        start_parameter="test",
        prices=[LabeledPrice(label="Zakaz", amount=100*100)]

    )


@fsm_add_adress.pre_checkout_query()
async def process_pre_checkout_query(pre_checkout_query:types.PreCheckoutQuery,bot:Bot)->None:

    await bot.answer_pre_checkout_query(pre_checkout_query.id,ok=True)

@fsm_add_adress.message(F.successful_payment)
async def process_successful_payment(message:types.Message,bot:Bot,state:FSMContext)->None:
    from private import start
    successful_payment_data = [
        message.from_user.id,
        message.successful_payment.total_amount / 100,
        message.successful_payment.currency,
        message.successful_payment.invoice_payload
    ]
    invoice_payload = successful_payment_data[3]
    try:
        await write_to_excel(successful_payment_data)

    except Exception as ex:
        print(ex)

    await message.answer("Покупка успешно завершена! Спасибо за ваш заказ.")
    data = await state.get_data()
    await bot.send_message(chat_id=ADMIN,text=f'Звать: {data["name"]}\nНомер: {data["phone"]}\nАдресс: {data["adress"]}\n\nТовары: {invoice_payload}')
    await state.clear()
    await start(message,state)

async def write_to_excel(successful_payment_data):
    import openpyxl
    from concurrent.futures import ThreadPoolExecutor
    import asyncio

    file_path = "success_payments.xlsx"

    def save_to_excel_sync():

        current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")


        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
            sheet = workbook.active
        else:

            workbook = openpyxl.Workbook()
            sheet = workbook.active

            sheet.append(["User ID", "Amount", "Currency", "Payload", "Date and Time"])


        sheet.append(successful_payment_data + [current_datetime])


        workbook.save(file_path)


    loop = asyncio.get_event_loop()
    with ThreadPoolExecutor() as pool:
        await loop.run_in_executor(pool, save_to_excel_sync)
